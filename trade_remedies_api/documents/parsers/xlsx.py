from io import BytesIO

import openpyxl
from openpyxl.chartsheet import Chartsheet


def parse(document):
    text = []
    workbook = openpyxl.load_workbook(filename=BytesIO(document.file.read()))
    document.file.close()
    for sheet in workbook.sheetnames:
        ws = workbook[sheet]
        if isinstance(ws, Chartsheet):
            # ChartSheets are sheets in Excel workbooks containing only a graph and nothing else,
            # they cause bugs when we attempt to parse them and so let's ignore them as their data
            # is not searchable anyway.
            continue
        col_range = ws[ws.min_column : ws.max_column]
        header_col_range = col_range[0]
        for cell in header_col_range:
            if cell_value := cell.value:
                text.append(cell_value)
    return "\n".join(text)
