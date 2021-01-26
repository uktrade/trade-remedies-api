import os
import tempfile
import pytest
from openpyxl import load_workbook
from core.exporters.writers.base_writer import BaseWriter
from core.exporters.writers.csv_writer import CSVWriter
from core.exporters.writers.excel_writer import ExcelWriter
from core.exporters import QuerysetExporter
from audit.models import Audit


@pytest.fixture
def writer_class_unimplemented():
    """Inadequately implemented writer class"""
    class UnimplementedWriter(BaseWriter):  # noqa
        def __init__(self, prefix):
            super().__init__("w+", prefix, ".txt")

    return UnimplementedWriter


@pytest.fixture
def writer_class_unimplemented_partial(writer_class_unimplemented):
    """Partially implemented writer class"""
    def fake(*args):  # noqa
        pass
    writer_class_unimplemented.setup = fake
    return writer_class_unimplemented


@pytest.fixture
def writer_class(mocker):
    """Concrete writer class"""
    class MyWriter(BaseWriter):
        def __init__(self, prefix):
            self.sensor = mocker.MagicMock()
            super().__init__("w+", prefix, ".txt")

        def setup(self):
            self.sensor.setup()

        def write_rows(self, rows):
            self.sensor.writerows(rows)

        def write_row(self, row):
            self.sensor.writerow(row)

        def close(self):
            self.sensor.close()
            return super().close()

    return MyWriter


@pytest.fixture
def row():
    """Export data item"""
    return ["hello", "exporter", "world"]


@pytest.fixture
def rows(row):
    """Export data items"""
    return [["1"] + row, ["2"] + row, ["3"] + row]


@pytest.fixture
def audits():
    """Set of audit records"""
    Audit.objects.create(type="CREATE").save()
    Audit.objects.create(type="UPDATE").save()
    Audit.objects.create(type="DELETE").save()
    Audit.objects.create(type="PURGE").save()
    Audit.objects.create(type="RESTORE").save()
    Audit.objects.create(type="READ").save()
    return Audit.objects.all().iterator()


@pytest.fixture
def batched_export(audits):
    """CSV Exporter with a custom batch size"""
    QuerysetExporter.BATCH_SIZE = 2

    def fake_writer(rows):
        assert len(rows) <= QuerysetExporter.BATCH_SIZE

    export = QuerysetExporter(queryset=audits, file_format="csv")
    export.writer.write_rows = fake_writer
    return export


class TestWriters:
    def test_writer_base_unimplemented(self, writer_class_unimplemented):
        with pytest.raises(NotImplementedError):
            # Without setup implemented, construction should raise NotImplementedError
            writer_class_unimplemented(prefix="foo")

    def test_writer_base_unimplemented_partial(self, writer_class_unimplemented_partial):
        writer = writer_class_unimplemented_partial(prefix="foo")
        with pytest.raises(NotImplementedError):
            writer.write_row([])
        with pytest.raises(NotImplementedError):
            writer.write_rows([])

    def test_writer_base_init(self, writer_class):
        writer = writer_class(prefix="foo")
        assert writer.sensor.setup.called
        assert writer.file
        # tempfile.NamedTemporaryFile returns a wrapper to manage file lifecycle
        assert isinstance(writer.file, tempfile._TemporaryFileWrapper)  # noqa
        assert os.stat(writer.file.name)
        assert writer.file.mode == "w+"
        assert "foo" in writer.file.name
        assert writer.file.name.endswith(".txt")

    def test_writer_base_cleanup(self, writer_class):
        writer = writer_class(prefix="foo")
        assert os.stat(writer.file.name)
        file_name = writer.close().name
        assert os.stat(file_name)
        del writer  # force writer (and tmp file refs) out of scope, i.e. cleanup
        with pytest.raises(FileNotFoundError):
            os.stat(file_name)

    def test_writer_base_write(self, writer_class, rows):
        writer = writer_class(prefix="foo")
        writer.write_row(rows[0])
        assert writer.sensor.write_row.called_with(rows[0])
        writer.write_rows(rows)
        assert writer.sensor.write_rows.called_with(rows)

    def test_writer_base_close(self, writer_class):
        writer = writer_class(prefix="foo")
        assert isinstance(writer.close(), tempfile._TemporaryFileWrapper)  # noqa
        assert writer.sensor.close_called

    def test_writer_csv(self, rows):
        writer = CSVWriter("foo")
        assert os.stat(writer.file.name)
        writer.write_row(rows[0])
        writer.write_rows(rows[1:])
        f = writer.close()
        f.seek(0)
        # un-csv the content and strip line endings to make comparison easier
        entries = [[i for i in map(lambda x: x.strip(), row.split(","))] for row in f.readlines()]
        assert entries == rows

    def test_writer_excel(self, rows):
        writer = ExcelWriter("foo")
        assert os.stat(writer.file.name)
        writer.write_row(rows[0])
        writer.write_rows(rows[1:])
        f = writer.close()
        # Load the excel export and iterate over the active worksheet
        wb = load_workbook(filename=f.name)
        ws = wb.active
        excel_data = []
        for item in ws.iter_rows():
            excel_data.append([item[0].value, item[1].value, item[2].value, item[3].value])
        assert excel_data == rows


class TestExporters:
    def test_export_format_unsupported(self):
        audits = Audit.objects.all().iterator()
        with pytest.raises(ValueError) as e:
            QuerysetExporter(queryset=audits, file_format="unsupported")
        assert "Unsupported export format: unsupported" in str(e)

    @pytest.mark.django_db
    def test_export_format_csv(self, audits):
        export = QuerysetExporter(queryset=audits, file_format="csv")
        export_file = export.do_export()
        export_file.seek(0)
        entries = [row for row in export_file.readlines()]
        assert len(entries) == 7  # 6 audits + header

    @pytest.mark.django_db
    def test_export_format_excel(self, audits):
        export = QuerysetExporter(queryset=audits, file_format="xlsx")
        export_file = export.do_export()
        wb = load_workbook(filename=export_file.name)
        ws = wb.active
        entries = [item for item in ws.iter_rows()]
        assert len(entries) == 7  # 6 audits + header

    @pytest.mark.django_db
    def test_export_queryset_invalid(self):
        # Use a QuerySet, not required GeneratorType
        audits = Audit.objects.all()
        with pytest.raises(ValueError) as e:
            QuerysetExporter(queryset=audits, file_format="xlsx")
        assert "queryset must be a generator" in str(e)

    @pytest.mark.django_db
    def test_export_queryset_empty(self):
        audits = Audit.objects.all().iterator()
        export = QuerysetExporter(queryset=audits, file_format="csv")
        export_file = export.do_export()
        export_file.seek(0)
        entries = [row for row in export_file.readlines()]
        assert entries[0] == "NO EXPORT DATA FOUND\n"

    @pytest.mark.django_db
    def test_export_compatible(self, audits):
        export = QuerysetExporter(queryset=audits, file_format="csv")
        export_file = export.do_export(compatible=True)
        export_file.seek(0)
        entry = export_file.readline()
        assert entry.strip("\n") == ",".join(Audit.row_columns())

    @pytest.mark.django_db
    def test_export_batch_size_generic(self, batched_export):
        batched_export.do_export()

    @pytest.mark.django_db
    def test_export_batch_size_compatible(self, batched_export):
        batched_export.do_export(compatible=True)
